# -*- coding: UTF-8 -*-
import csv
import sys
import numpy as np
import copy
import math
Transitions = {}
Cohort = {}
Reward = {}
c=[2,3,6,3,2.5]
#user_act=1 #0:sleep 1:watch TV
sensor_number=[1,1,1,1,1]
r=1-0.9 #1-r:0,0.1,0.2,0.3
d=0.1
if len(sys.argv)>2:
    r = 1- float(sys.argv[2])
r=round(r, 1) 
num_iteration=20
reliable=True#True:method B ; False : method A
all_sensor_on=True
save_results=False
all_sensor_number=np.sum(sensor_number)
#gamma is the discount factor
if len(sys.argv)>1:
    gamma = float(sys.argv[1])
else:
    gamma = 0.9
sleep_sensor=sensor_number[1]+sensor_number[2]+sensor_number[3]
wTV_sensor=sensor_number[4]+sensor_number[3]+sensor_number[1]
#the maximum error allowed in the utility of any state
if len(sys.argv)>4:
    epsilon = float(sys.argv[4])
else:
    epsilon = 0.001

def read_MDP():
	val=['nor','on_','off']
	all_states=[]
	for i in range(3**all_sensor_number):
		v=i
		state=[]
		for j in range(all_sensor_number):
			state.append(v%3)
			v=int(v/3)
		all_states.append(state)
	#print(all_states)

	for ptr_state in all_states:
		act={}
		Nact={}
		#######
		#sleep#
		#######
		m1=[0]*3 #0:'sleep' 1:'-sleep' 2:'nor'
		m2=[0]*3
		m_wTV=[0]*3
		

		for i in range(sensor_number[2-1]):
			if ptr_state[sensor_number[0]+i]==1:
				  #night
				m2[0]+=0.3*(1-d)
				m2[2]+=1-0.3*(1-d)
			elif ptr_state[sensor_number[0]+i]==2:

				m2[1]+=0.3*(1-d)
				m2[2]+=1-0.3*(1-d)
			else:
				m2[2]+=1

		for i in range(sensor_number[3-1]):
			if ptr_state[sensor_number[0]+sensor_number[1]+i]==1:
				#bed
				m2[0]+=1*(1-d)
			elif ptr_state[sensor_number[0]+sensor_number[1]+i]==2:
				m2[1]+=1*(1-d)
			else:
				m2[2]+=1

		for i in range(sensor_number[4-1]):
			if ptr_state[sensor_number[0]+sensor_number[1]+sensor_number[2]+i]==1:
				  #card
				m2[0]+=1*(1-d)
			elif ptr_state[sensor_number[0]+sensor_number[1]+sensor_number[2]+i]==2:
				
				m2[1]+=1*(1-d)
			else:
				m2[2]+=1

		m2[0]/=3
		m2[1]/=3
		m2[2]/=3
		for i in range(sensor_number[1-1]):
			if ptr_state[i]==1:
				#clock
				m1[0]=0.8*(1-d)
				m1[2]=1-0.8*(1-d)
			elif ptr_state[i]==2:
				m1[1]=0.8*(1-d)
				m1[2]=1-0.8*(1-d)
			else:
				m1[2]=1
		U=[set({'sleep'}),set({'-sleep'}),set({'sleep','-sleep'})]
		m=0
		m_Den=0
		#print(m1,m2)
		for clock in range(len(U)):
			for NBC in range(len(U)):
				m_Den+=m1[clock]*m2[NBC]
				if U[clock]&U[NBC]==set({'sleep'}):
					m+=m1[clock]*m2[NBC]
		
		if m_Den!=0:
			belief_sleep= m/m_Den
		else:
			belief_sleep=0
		#########
		#watchTV#
		#########
		
		m_wTV[0]=0
		m_wTV[1]=0
		m_wTV[2]=0
		for i in range(sensor_number[5-1]):
			if ptr_state[sensor_number[0]+sensor_number[1]+sensor_number[2]+sensor_number[3]+i]==1:
				#TV
				m_wTV[0]+=1
			elif ptr_state[sensor_number[0]+sensor_number[1]+sensor_number[2]+sensor_number[3]+i]==2:
				m_wTV[1]+=1
			else:
				m_wTV[2]+=1
				

		for i in range(sensor_number[4-1]):
			if ptr_state[sensor_number[0]+sensor_number[1]+sensor_number[2]+i]==1:
				  #card
				m_wTV[0]+=1
			elif ptr_state[sensor_number[0]+sensor_number[1]+sensor_number[2]+i]==2:
				m_wTV[1]+=1
			else:
				m_wTV[2]+=1
				

		for i in range(sensor_number[2-1]):
			if ptr_state[sensor_number[0]+i]==1:
				#day
				m_wTV[0]+=0.7
				m_wTV[2]+=0.3
			elif ptr_state[sensor_number[0]+i]==2:
				m_wTV[1]+=0.7
				m_wTV[2]+=0.3
			else:
				m_wTV[2]+=1
		m_wTV[0]/=3
		m_wTV[1]/=3
		m_wTV[2]/=3

		
		belief_wTV=m_wTV[0]
		#print(ptr_state)
		#print(s2,s3,s4,s5,s6,s7)
		#print(belief_sleep,belief_wTV)
		#choose Turn on sensor action
		state_cost=0
		tag=0
		for t in range(5):
			for i in range(sensor_number[t]):
				
				if ptr_state[tag]!=0:
					state_cost+=c[t]
				tag+=1

		state_error=(np.dot(c,sensor_number)-state_cost)/np.dot(c,sensor_number)

		#state_error*=2
		#w1=0.5
		#w2=1-w1
		belief=0
		o_sensor=[]
		wTV_LB=0.4
		sleep_LB=0.4
		UB=1
		min_LB=min(wTV_LB,sleep_LB)
		if belief_sleep>belief_wTV :
			belief=belief_sleep
			if belief <sleep_LB:
				belief =0
			# else:
				# (belief-sleep_LB)/(UB-sleep_LB)
			o_sensor=[0,1,2,3]
		elif(belief_sleep<belief_wTV ):
			belief=belief_wTV
			if belief <wTV_LB:
				belief =0
			# else:
				# (belief-wTV_LB)/(UB-wTV_LB)
			o_sensor=[4,3,1]
		else:
			belief=belief_sleep
			if belief <min_LB:
				belief =0
			# else:
				# (belief-min_LB)/(UB-min_LB)
			o_sensor=[0,1,2,3,4]
		if belief>UB:
			belief=UB
			
		tag=0
		for t in range(5):
			for i in range(sensor_number[t]):
				if ptr_state[tag]==0 and t in o_sensor:
					on=ptr_state[:]
					off=ptr_state[:]
					on[tag]=1
					off[tag]=2
					p_act='Turn on Sensor'+str(tag+1)
					act[p_act]=[(0.5,"("+" ".join(str(ele) for ele in on)+")"),(0.5,"("+" ".join(str(ele) for ele in off)+")")]
					Nact[p_act]=[[1,"("+" ".join(str(ele) for ele in on)+")"],[1,"("+" ".join(str(ele) for ele in off)+")"]]
				tag+=1
		#print(act)
		#state reward
		

		#choose Turn off sensor action
		if belief >0:
			tag=0
			for t in range(5):
				for i in range(sensor_number[t]):
					if ptr_state[tag]!=0:
						nor=ptr_state[:]
						nor[tag]=0
						p_act='Turn off Sensor'+str(tag+1)
						act[p_act]=[(1,"("+" ".join(str(ele) for ele in nor)+")")]
						Nact[p_act]=[[1,"("+" ".join(str(ele) for ele in nor)+")"]]
					tag+=1


		#state reward
		p_state="("+" ".join(str(ele) for ele in ptr_state)+")"
		#Reward[p_state]=((state_error*0.13)+(0.87*belief)) if state_error*belief!=0 else 0
		Wer=0.2
		#Reward[p_state]=(Wer*state_error+belief*(1-Wer))*state_error*belief
		#print(p_state)
		#e_x=math.exp(state_error*1)
		#e_negax=math.exp(-state_error*1)
		#Reward[p_state]=(state_error)*(belief)
		#Reward[p_state]=(state_error)*((e_x-e_negax)/(e_x+e_negax))
		#Reward[p_state]=(state_error)*(1/(1+math.exp(-(belief*20-10))))
		Reward[p_state]=(Wer*state_error+belief*(1-Wer)) if belief!=0 else 0
		#print(Reward[p_state])
		Reward["(0 1 0 1 0)"]=0
		Reward["(0 1 2 1 0)"]=0
		Reward["(0 1 0 1 2)"]=0
		Reward["(0 1 2 1 2)"]=0
		Reward["(2 1 2 1 0)"]=0
		Reward["(2 1 0 1 2)"]=0
		Reward["(2 1 2 1 2)"]=0
		Reward["(2 1 0 1 0)"]=0



		#keep action

		turn_off_sensor=[]
		tag=0
		for t in range(5):
			for i in range(sensor_number[t]):
				if ptr_state[tag] ==0:
					turn_off_sensor.append(tag)
				tag+=1
		p = 1 if len(turn_off_sensor)<all_sensor_number else 1
		p_act='keep'
		act[p_act]=[(p,p_state)]
		Nact[p_act]=[[1,p_state]]
		if p!=1:
			if len(turn_off_sensor)!=all_sensor_number:
				for state in all_states:
					turn_off=[]
					for t in range(all_sensor_number):
						if state[t] ==0:
							turn_off.append(t)
					#s1=set(turn_off_sensor)
					#s2=set(turn_off)
					if turn_off_sensor==turn_off :
						if ptr_state!=state:
							act[p_act].append(((1-p)/(2**(all_sensor_number-len(turn_off))-1),"("+" ".join(str(ele) for ele in state)+")"))
		

					
		Transitions[p_state]=act
		Cohort[p_state]=Nact
		#print("("+" ".join(str(ele) for ele in ptr_state)+")")
		


read_MDP()
#print(Cohort['(0 0 0 0 0)']['keep'][0][1])

class MarkovDecisionProcess:

    """A Markov Decision Process, defined by an states, actions, transition model and reward function."""

    def __init__(self, transition={}, reward={}, gamma=.9):
        #collect all nodes from the transition models
        self.states = transition.keys()
        #initialize transition
        self.transition = transition
        #initialize reward
        self.reward = reward
        #initialize gamma
        self.gamma = gamma

    def R(self, state):
        """return reward for this state."""
        return self.reward[state]

    def actions(self, state):
        """return set of actions that can be performed in this state"""
        return self.transition[state].keys()

    def T(self, state, action):
        """for a state and an action, return a list of (probability, result-state) pairs."""
        return self.transition[state][action]

#Initialize the MarkovDecisionProcess object
mdp = MarkovDecisionProcess(transition=Transitions, reward=Reward)

def value_iteration(mdp):
    """
    Solving the MDP by value iteration.
    returns utility values for states after convergence
    """
    states = mdp.states
    actions = mdp.actions
    T = mdp.T
    R = mdp.R
    #initialize value of all the states to 0 (this is k=0 case)
    V1 = {s: 0 for s in states}
    while True:
        V = V1.copy()
        delta = 0
        for s in states:
            #Bellman update, update the utility values
            V1[s] = R(s) + gamma * max([ sum([p * V[s1] for (p, s1) in T(s, a)]) for a in actions(s)])
            #V1[s] = R(s) + gamma * max([ (sum([p * V[s1] for (p, s1) in T(s, a)])-1 if "Turn" in a else sum([p * V[s1] for (p, s1) in T(s, a)]) ) for a in actions(s)])
            #calculate maximum difference in value
            delta = max(delta, abs(V1[s] - V[s]))

        #check for convergence, if values converged then return V
        if delta < epsilon * (1 - gamma) / gamma:
            return V


def best_policy(V):
    """
    Given an MDP and a utility values V, determine the best policy as a mapping from state to action.
    returns policies which is dictionary of the form {state1: action1, state2: action2}
    """
    states = mdp.states
    actions = mdp.actions
    pi = {}
    for s in states:
        pi[s] = max(actions(s), key=lambda a: expected_utility(a, s, V))
    return pi


def expected_utility(a, s, V):
    """returns the expected utility of doing a in state s, according to the MDP and V."""
    T = mdp.T
    return sum([p * V[s1] for (p, s1) in mdp.T(s, a)])

import random

#sensor_r=range(5)
#for i in range(5):
#	sensor_r[i]=random.choice([0.1,0.15,0.2,0.3])
curr_state=""
def set_init_state():
	curr_state="("
	for i in range(all_sensor_number):
		curr_state=curr_state+"0"
		if i!=all_sensor_number-1:
			curr_state=curr_state+" "
	curr_state=curr_state+")"
	return curr_state
#(0 0 .... 0 0) start
curr_state=set_init_state()
if all_sensor_on:
	curr_state="(1 1 1 1 1)"
#print(curr_state)
#save pi
#pin=np.array(pi)
#np.save('pi.npy',pin)

# def env(state,action):
	# s=np.zeros(5)
	# invert=2
	# pick_sensor=1
	# if(user_act==0):#sleep
		# s[0]=random.randint(1, 2)
		# s[1]=1
		# s[2]=1
		# s[3]=1
		# s[4]=2
	# else:#wTV
		# s[4]=1
		# s[3]=1
		# s[1]=1
		# s[0]=2
		# s[2]=2
	# if "Turn" in action:
		# pick_sensor=int(action[-1])
		# invert=1 if s[pick_sensor-1]==2 else 2
	
	# if "Turn on" in action:
		# state=state[:pick_sensor*2-1]+"1"+state[pick_sensor*2:]#1 3 5 7 9
		# for j in range(len(state)):
			# i=(j-1)/2
			# if state[j]=="1" or state[j]=="2":
				# invert=1 if s[i]==2 else 2
				# state=state[:j]+str(int(np.random.choice([s[i],invert],size=1,p=[1-sensor_r[i],sensor_r[i]])[0]))+state[j+1:]
	# elif "Turn off" in action:
		# state=state[:pick_sensor*2-1]+"0"+state[pick_sensor*2:]
		# for j in range(len(state)):
			# i=(j-1)/2
			# if state[j]=="1" or state[j]=="2":
				# invert=1 if s[i]==2 else 2
				# state=state[:j]+str(int(np.random.choice([s[i],invert],size=1,p=[1-sensor_r[i],sensor_r[i]])[0]))+state[j+1:]
	# elif "keep" in action:
		# for j in range(len(state)):
			# i=(j-1)/2
			# if state[j]=="1" or state[j]=="2":
				# invert=1 if s[i]==2 else 2
				# state=state[:j]+str(int(np.random.choice([s[i],invert],size=1,p=[1-sensor_r[i],sensor_r[i]])[0]))+state[j+1:]
	# return state



# f = open('./data/value/value train.txt','r')
# real_sensor=f.readline()
# fla = open('./data/label/label train.txt','r')
# label=fla.readline()
def env(state,action,real_sensor):

	
	#print(real_sensor)
	pick_sensor=1
	turn_on_off_sensor=False
	if all_sensor_on:
		next_state="(1 1 1 1 1)"
	else:
		if "Turn" in action:
			pick_sensor=int(action[-1])
			turn_on_off_sensor=True
		if "Turn on" in action:
			next_state=state[:pick_sensor*2-1]+"1"+state[pick_sensor*2:]#1 3 5 7 9

		elif "Turn off" in action:
			next_state=state[:pick_sensor*2-1]+"0"+state[pick_sensor*2:]
		else:
			next_state=state
	#print(real_sensor,next_state)
	for j in range(len(next_state)):
		if next_state[j]=="1" or next_state[j]=="2":
			next_state=next_state[:j]+real_sensor[j-1]+next_state[j+1:]
#	elif "keep" in action:

	#print(state,next_state)
	#print(state,action)
	return next_state,turn_on_off_sensor
	
	
def cal_belief(curr_state):
		#print("Phase : "+str(t)+"  state : "+curr_state+"   action : "+pi[curr_state])
	belief_sleep=0
	belief_wTV=0

	#######
	#sleep#
	#######
	m1=[0]*3 #0:'sleep' 1:'-sleep' 2:'nor'
	m2=[0]*3
	m_wTV=[0]*3
	

	
	if curr_state[1*2+1]=='1':
		  #night
		m2[0]+=0.3*(1-d)
		m2[2]+=1-0.3*(1-d)
	elif curr_state[1*2+1]=='2':

		m2[1]+=0.3*(1-d)
		m2[2]+=1-0.3*(1-d)
	else:
		m2[2]+=1
	

	
	if curr_state[2*2+1]=='1':
		#bed
		m2[0]+=1*(1-d)
	elif curr_state[2*2+1]=='2':

		m2[1]+=1*(1-d)
	else:
		m2[2]+=1

	
	if curr_state[3*2+1]=='1':
		  #card
		m2[0]+=1*(1-d)
	elif curr_state[3*2+1]=='2':
		
		m2[1]+=1*(1-d)
	else:
		m2[2]+=1
	m2[0]/=3
	m2[1]/=3
	m2[2]/=3

	if curr_state[0*2+1]=='1':
		#clock
		m1[0]=0.8*(1-d)
		m1[2]=1-0.8*(1-d)
	elif curr_state[0*2+1]=='2':
		m1[1]=0.8*(1-d)
		m1[2]=1-0.8*(1-d)
	else:
		m1[2]=1
	U=[set({'sleep'}),set({'-sleep'}),set({'sleep','-sleep'})]
	m=0
	m_Den=0
	for clock in range(len(U)):
		for NBC in range(len(U)):
			m_Den+=m1[clock]*m2[NBC]
			if U[clock]&U[NBC]==set({'sleep'}):
				m+=m1[clock]*m2[NBC]
	
	if m_Den!=0:
		belief_sleep= m/m_Den
	else:
		belief_sleep=0
	#########
	#watchTV#
	#########
	
	m_wTV[0]=0
	m_wTV[1]=0
	m_wTV[2]=0

	if curr_state[4*2+1]=='1':
		#TV
		m_wTV[0]+=1*(1-d)
	elif curr_state[4*2+1]=='2':
		m_wTV[1]+=1*(1-d)
	else:
		m_wTV[2]+=1
			


	if curr_state[3*2+1]=='1':
		  #card
		m_wTV[0]+=1*(1-d)
	elif curr_state[3*2+1]=='2':
		m_wTV[1]+=1*(1-d)
	else:
		m_wTV[2]+=1
			


	if curr_state[1*2+1]=='1':
		#day
		m_wTV[0]+=0.7*(1-d)
		m_wTV[2]+=1-0.7*(1-d)
	elif curr_state[1*2+1]=='2':
		m_wTV[1]+=0.7*(1-d)
		m_wTV[2]+=1-0.7*(1-d)
	else:
		m_wTV[2]+=1
	m_wTV[0]/=3
	m_wTV[1]/=3
	m_wTV[2]/=3

	
	belief_wTV=m_wTV[0]
	
	#print('識別居民行為 :'),
	# if belief_sleep<belief_wTV and belief_wTV>0.4:
		# #print("watch TV")
		# if "watch TV"in label:
			# hit+=1

	# elif(belief_sleep>0.4):
		# #print("sleep")
		# if "sleep"in label:
			# hit+=1
	# else:
		# print("uncertain")
	return belief_sleep,belief_wTV

def save_excel(Accuracy_list,Energy_list,ws):
	
	ws.insert_cols(2, 2)
	ws.cell(column=2, row=1).value = 'Accuracy(gamma='+str(gamma)+')'
	ws.cell(column=3, row=1).value = 'Energy Rate(gamma='+str(gamma)+')'
	for t in range(num_iteration):
		ws.cell(column=1, row=t+2).value = t+1
		ws.cell(column=2, row=t+2).value =Accuracy_list[t]
		ws.cell(column=3, row=t+2).value =Energy_list[t]
def update_Cohort_Table(Cohort):
	Transitions=copy.deepcopy(Cohort)
	for s in Cohort:
		for a in Cohort[s]:
			count=0
			for ns_list in Cohort[s][a]:
				count+=ns_list[0]
			for ns in range(len(Cohort[s][a])):
				Transitions[s][a][ns]=(Transitions[s][a][ns][0]/count,Cohort[s][a][ns][1])
	return Transitions

# def test():
	# curr_state=set_init_state()
	# # f = open('./data/value/value test.txt','r')
	# # real_sensor=f.readline()
	# fla = open('./data/label/label r '+str(r)+'.txt','r')
	# label=fla.readline()
	# count=0
	# hit=0
	# all_en=0
	# pWaW=0
	# pWaS=0
	# pSaS=0
	# pSaW=0
	# for	t in range(0,86400):
		# #print(curr_state)
		# belief_sleep,belief_wTV=cal_belief(curr_state)
		# if reliable:

			# if curr_state!="(0 1 0 1 0)" and curr_state!="(0 1 2 1 0)" and curr_state!="(0 1 0 1 2)" and curr_state!="(0 1 2 1 2)"and curr_state!="(2 1 2 1 2)"and curr_state!="(2 1 0 1 0)"and curr_state!="(2 1 2 1 0)"and curr_state!="(2 1 0 1 2)":

				# if belief_sleep<belief_wTV and belief_wTV>0.4:
					# count+=1
					# if "watch TV"in label:
						# hit+=1
						# pWaW+=1
						# #print("watch TV")
					# else:
						# pWaS+=1
						# #print(label)
				# elif(belief_sleep>0.4):
					# count+=1
					# #print("sleep")
					# if "sleep"in label:
						# hit+=1
						# pSaS+=1
						# #print("sleep")
					# else:
						# pSaW+=1
						# #print(label)

		# else:
			# if belief_sleep<belief_wTV :
				# #print("watch TV")
				# count+=1
				# if "watch TV"in label:
					# hit+=1
					# pWaW+=1
					# #print("watch TV")
				# else:
					# pWaS+=1
					# #print(label)

			# else:
				# #print("sleep")
				# count+=1
				# if "sleep"in label:
					# hit+=1
					# pSaS+=1
					# #print("sleep")
				# else:
					# pSaW+=1
					# #print(label)
			# #print(curr_state)
		
		# en=0
		# tag=0
		# # for i in range(5):
			# # for j in range(sensor_number[i]):
				# # if curr_state[tag*2+1]=='1' or curr_state[tag*2+1]=='2':
					# # en+=c[i]
				# # tag+=1
		# for i in range(5):
			# if curr_state[i*2+1]=='1' or curr_state[i*2+1]=='2':
				# en+=c[i]
		# all_en+=en
		
		# # if t%86398==0:
			# # update_Cohort_Table()
			# # mdp = MarkovDecisionProcess(transition=Transitions, reward=Reward)
			# # V = value_iteration()
			# # pi = best_policy(V)
		# if t!=86399:
			# curr_state,_=env(curr_state,pi[curr_state])	

		# label=fla.readline()
		# #real_sensor=f.readline()

		# #print(label)
	# print("辨識次數:",count)
	
	# print("Accuracy:",hit/count)
	# print("能耗合:",all_en)	
	# #print(all_en/(count*np.dot(c,sensor_number)))
	# if all_sensor_on:
		# print("耗能率:",all_en/(86399*np.sum(c)))
	# else:
		# print("耗能率:",all_en/(count*np.sum(c)))

	# print('     W    S   <actual class')
	# print('W',pWaW,pWaS)
	# print('S',pSaW,pSaS)
	# print('^ predicted class')
	# Precision_W=pWaW/(pWaW+pSaW)
	# Recall_W=pWaW/(pWaW+pWaS)
	# Precision_S=pSaS/(pSaS+pWaS)
	# Recall_S=pSaS/(pSaS+pSaW)
	# print('Precision_Watch TV:',Precision_W)
	# print('Recall_Watch TV:',Recall_W)
	# print('Precision_Sleep:',Precision_S)
	# print('Recall_Sleep:',Recall_S)
	#call value iteration
def train():
	curr_state=set_init_state()
	Accuracy_list=[]
	Energy_list=[]
	print("r=",1-r)
	for e in range(num_iteration):
		num_turn_on_off_sensor=0
		print(e+1)
		f = open('./data/value/value r '+str(r)+'.txt','r')
		real_sensor=f.readline()
		fla = open('./data/label/label r '+str(r)+'.txt','r')
		label=fla.readline()
		count=0
		hit=0
		all_en=0
		pWaW=0
		pWaS=0
		pSaS=0
		pSaW=0
		for	t in range(0,86400):
			#print(curr_state)
			if t==0:
				if not all_sensor_on:
					Transitions=update_Cohort_Table(Cohort)
					mdp = MarkovDecisionProcess(transition=Transitions, reward=Reward)
					V = value_iteration(mdp)
					pi = best_policy(V)
			if not all_sensor_on:		
				action=pi[curr_state]
			else:
				action='keep'
			#print(curr_state)
			belief_sleep,belief_wTV=cal_belief(curr_state)
			if reliable:

				if curr_state!="(0 1 0 1 0)" and curr_state!="(0 1 2 1 0)" and curr_state!="(0 1 0 1 2)" and curr_state!="(0 1 2 1 2)"and curr_state!="(2 1 2 1 2)"and curr_state!="(2 1 0 1 0)"and curr_state!="(2 1 2 1 0)"and curr_state!="(2 1 0 1 2)":

					if belief_sleep<belief_wTV and belief_wTV>0.4:
						count+=1
						if "watch TV"in label:
							hit+=1
							pWaW+=1
							#print("watch TV")
						else:
							pWaS+=1
							# print('===============')
							# print(curr_state)
							# print(label)
							# print(belief_sleep,belief_wTV)
							# print(action)
					elif(belief_sleep>0.4):
						count+=1
						#print("sleep")
						if "sleep"in label:
							hit+=1
							pSaS+=1
							#print("sleep")
						else:
							pSaW+=1
							# print('===============')
							# print(curr_state)
							# print(label)
							# print(belief_sleep,belief_wTV)
							# print(action)

			else:
				if belief_sleep<belief_wTV :
					#print("watch TV")
					count+=1
					if "watch TV"in label:
						hit+=1
						pWaW+=1
						#print("watch TV")
					else:
						pWaS+=1
						#print(label)

				else:
					#print("sleep")
					count+=1
					if "sleep"in label:
						hit+=1
						pSaS+=1
						#print("sleep")
					else:
						pSaW+=1
						#print(label)
				#print(curr_state)
			
			en=0
			tag=0
			# for i in range(5):
				# for j in range(sensor_number[i]):
					# if curr_state[tag*2+1]=='1' or curr_state[tag*2+1]=='2':
						# en+=c[i]
					# tag+=1
			for i in range(5):
				if curr_state[i*2+1]=='1' or curr_state[i*2+1]=='2':
					en+=c[i]
			all_en+=en
			if t!=86399:
				real_sensor=f.readline()
				#print(real_sensor)
				
				next_state,turn_on_or_off_sensor=env(curr_state,action,real_sensor)	
			if not all_sensor_on:
				flag=False
				for i in range(len(Cohort[curr_state][action])):
					if Cohort[curr_state][action][i][1]==next_state:
						#print(Cohort[state][action][i])
						Cohort[curr_state][action][i][0]+=1
						flag=True
				if flag==False:
					Cohort[curr_state][action].append([1,next_state])
			curr_state=next_state
			label=fla.readline()
			if turn_on_or_off_sensor:
				num_turn_on_off_sensor+=1
		#print("num of turn on or off sensor:",num_turn_on_off_sensor)
		print("辨識次數:",count)
		Accuracy=hit/count

		if pSaS!=0:
			Precision_S=pSaS/(pSaS+pWaS)
			Recall_S=pSaS/(pSaS+pSaW)
			beta=1
			F1_S=(1+beta**2)*(Precision_S*Recall_S)/(beta**2*Precision_S+Recall_S)
		else:
			Recall_S=0
			beta=1
			F1_S=0

		if pWaW!=0:
			Precision_W=pWaW/(pWaW+pSaW)
			Recall_W=pWaW/(pWaW+pWaS)
			beta=1
			F1_W=(1+beta**2)*(Precision_W*Recall_W)/(beta**2*Precision_W+Recall_W)
		else:
			Recall_W=0
			beta=1
			F1_W=0

		print('F1-score_Sleep:',F1_S)
		print('F1-score_W_TV:',F1_W)
		print("Accuracy:",Accuracy)
		print("能耗合:",all_en)	
		#print(all_en/(count*np.dot(c,sensor_number)))
		Energy =all_en/(86400*np.sum(c))
		print("耗能率:",Energy)
		# if (e+1)%1==0:
			# f = open('./data/value/value r '+str(r)+'.txt','r')
			# real_sensor=f.readline()
			# test()
		Accuracy_list.append(round(Accuracy,4))
		Energy_list.append(round(Energy,4))
	return Accuracy_list,Energy_list
if __name__ == '__main__':
	if save_results:
		init_Cohort=copy.deepcopy(Cohort)
		import openpyxl
		fn = str(1-r)+'.xlsx'
		
		wb = openpyxl.Workbook()
		ws=wb.active
		ws.cell(column=1, row=1).value = 'iteration'
		for i in range(9):
			gamma=round((i+1)*0.1,1)
			print('gamma='+str(gamma))
			Accuracy_list,Energy_list=train()
			Cohort=copy.deepcopy(init_Cohort)
			
			save_excel(Accuracy_list,Energy_list,ws)
		wb.save(fn)
	else:
		train()
#call value iteration
"""

print '    State    -     Value'
for s in V:
    print s, ' - ' , V[s]
pi = best_policy(V)
print '\nOptimal policy is \n   State     -     Action'
for s in pi:
    print s, ' - ' , pi[s]
"""
